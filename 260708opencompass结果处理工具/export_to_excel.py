#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Export OpenCompass evaluation outputs to a clean Excel workbook.

Usage:
  python tools/export_to_excel.py --input <run_folder> --output <output.xlsx>

Defaults:
  --input  omitted: use the newest folder under outputs/default
  --output omitted: create <model>-<dataset>-<run_timestamp>.xlsx under the input folder

Workbook sheets:
  1. summary: content parsed from summary/*.md
  2. details: 题号 / 是否正确 / 模型结果 / 原始题目 / 标准答案 / 回答是否为空
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit("缺少 openpyxl：请先运行 pip install openpyxl") from exc


TRUE_VALUES = {"true", "1", "yes", "y", "是", "正确"}
QUESTION_MARKERS = ["Question:", "问题：", "问题:", "题目：", "题目:", "Q:"]
ANSWER_CUT_PATTERNS = [
    r"\n\s*Let's think step by step\s*\n?",
    r"\n\s*Answer\s*:\s*",
    r"\n\s*答案\s*[:：]\s*",
    r"\n\s*参考答案\s*[:：]\s*",
]


def die(message: str) -> None:
    raise SystemExit(f"[export_to_excel] {message}")


def read_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def latest_dir(base: Path) -> Path:
    if not base.exists():
        die(f"默认目录不存在：{base}")

    # OpenCompass 可能是 outputs/default/<run_timestamp>，
    # 也可能是 outputs/default/<exp_name>/<run_timestamp>。
    # 所以递归找真正含有 summary 和 results 的运行目录。
    candidates = []
    for p in base.rglob("*"):
        if not p.is_dir():
            continue
        if (p / "summary").exists() and (p / "results").exists():
            candidates.append(p)

    if not candidates:
        die(f"默认目录下没有找到包含 summary/ 和 results/ 的结果文件夹：{base}")

    # 文件夹名里通常带时间戳；mtime 可兜底。
    return sorted(candidates, key=lambda p: (p.name, p.stat().st_mtime))[-1]


def latest_file(paths: Iterable[Path]) -> Optional[Path]:
    files = [p for p in paths if p.is_file()]
    if not files:
        return None
    return sorted(files, key=lambda p: (p.name, p.stat().st_mtime))[-1]


def find_summary_md(input_dir: Path) -> Path:
    path = latest_file((input_dir / "summary").glob("*.md"))
    if path is None:
        path = latest_file(input_dir.rglob("summary*.md"))
    if path is None:
        die(f"没有找到 summary md：{input_dir}/summary/*.md")
    return path


def find_result_files(input_dir: Path) -> List[Path]:
    results_dir = input_dir / "results"
    files = sorted(results_dir.rglob("*.json")) if results_dir.exists() else []
    if not files:
        die(f"没有找到 result json：{results_dir}/**/*.json")
    # 排除明显不是逐题结果的 json 可在这里加规则；目前保留全部 dataset json。
    return files


def find_prediction_file(input_dir: Path, result_file: Path) -> Optional[Path]:
    pred_dir = input_dir / "predictions"
    if not pred_dir.exists():
        return None

    candidates = list(pred_dir.rglob(result_file.name))
    if candidates:
        return latest_file(candidates)

    # 文件名不完全相同时，按 stem 模糊找。
    candidates = list(pred_dir.rglob(f"*{result_file.stem}*.json"))
    return latest_file(candidates)


def scalar(value: Any) -> Any:
    """OpenCompass 里 pred/answer/correct 常是单元素 list，这里压成普通值。"""
    if isinstance(value, list):
        if not value:
            return ""
        if len(value) == 1:
            return scalar(value[0])
        return " | ".join(str(scalar(v)) for v in value)
    if value is None:
        return ""
    return value


def to_bool(value: Any) -> bool:
    value = scalar(value)
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in TRUE_VALUES


def is_empty_answer(value: Any) -> bool:
    value = scalar(value)
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return str(value).strip() == ""


def normalize_result_details(result_json: Any) -> List[Dict[str, Any]]:
    if isinstance(result_json, dict):
        if isinstance(result_json.get("details"), list):
            return result_json["details"]
        # 有些结果可能直接是 {id: detail}
        if all(isinstance(v, dict) for v in result_json.values()):
            return list(result_json.values())
    if isinstance(result_json, list):
        return [x for x in result_json if isinstance(x, dict)]
    die("result json 格式无法识别：没有 details 列表")


def normalize_prediction_map(pred_json: Any) -> Tuple[Dict[str, Dict[str, Any]], List[Dict[str, Any]]]:
    """返回 index/key -> prediction record，以及按顺序的 list。"""
    by_key: Dict[str, Dict[str, Any]] = {}
    rows: List[Dict[str, Any]] = []

    if pred_json is None:
        return by_key, rows

    if isinstance(pred_json, dict):
        for k, v in pred_json.items():
            if isinstance(v, dict):
                by_key[str(k)] = v
                rows.append(v)
    elif isinstance(pred_json, list):
        for i, v in enumerate(pred_json):
            if isinstance(v, dict):
                by_key[str(i)] = v
                rows.append(v)
    return by_key, rows


def index_from_abbr(example_abbr: Any) -> Optional[str]:
    text = str(example_abbr or "")
    m = re.search(r"(?:^|_)(\d+)$", text)
    return m.group(1) if m else None


def pick_prediction_record(
    detail: Dict[str, Any], i: int, pred_by_key: Dict[str, Dict[str, Any]], pred_rows: List[Dict[str, Any]]
) -> Dict[str, Any]:
    abbr = str(detail.get("example_abbr", ""))
    idx = index_from_abbr(abbr)

    for key in (abbr, str(i), idx):
        if key is not None and key in pred_by_key:
            return pred_by_key[key]

    if i < len(pred_rows):
        return pred_rows[i]
    return {}


def extract_current_question(origin_prompt: Any) -> str:
    """从 few-shot prompt 中只抽最后一道题，去掉示例和 Answer 引导。"""
    text = str(origin_prompt or "").replace("\r\n", "\n")
    if not text.strip():
        return ""

    # 找最后一个题目标记；GSM8K few-shot 里每题都以 Question: 开头。
    last_pos = -1
    last_marker = ""
    for marker in QUESTION_MARKERS:
        pos = text.rfind(marker)
        if pos > last_pos:
            last_pos = pos
            last_marker = marker

    if last_pos >= 0:
        text = text[last_pos + len(last_marker):]

    # 截掉答案触发语，但保留选项等题干内容。
    cut_positions = []
    for pattern in ANSWER_CUT_PATTERNS:
        m = re.search(pattern, text, flags=re.IGNORECASE)
        if m:
            cut_positions.append(m.start())
    if cut_positions:
        text = text[: min(cut_positions)]

    return text.strip()


def extract_dataset_name(result_file: Path, result_json: Any) -> str:
    if isinstance(result_json, dict):
        details = result_json.get("details")
        if isinstance(details, list) and details:
            abbr = str(details[0].get("example_abbr", ""))
            m = re.match(r"(.+?)_(?:test|val|dev|train)_\d+$", abbr)
            if m:
                return m.group(1)
    return result_file.stem


def parse_markdown_table(md: str) -> Optional[List[List[str]]]:
    lines = [line.strip() for line in md.splitlines() if line.strip()]
    table_lines = [line for line in lines if line.startswith("|") and line.endswith("|")]
    if not table_lines:
        return None

    rows: List[List[str]] = []
    for line in table_lines:
        cells = [c.strip() for c in line.strip("|").split("|")]
        # 跳过 markdown 分隔行：| --- | --- |
        if cells and all(re.fullmatch(r":?-{3,}:?", c.replace(" ", "")) for c in cells):
            continue
        rows.append(cells)
    return rows or None


def infer_default_output_name(summary_md: str, input_dir: Path, datasets: List[str]) -> str:
    table = parse_markdown_table(summary_md)
    model = "model"
    dataset = "+".join(sorted(set(datasets))) if datasets else "dataset"

    if table and len(table) >= 2:
        header = table[0]
        first_row = table[1]
        if first_row:
            dataset = "+".join(sorted(set(row[0] for row in table[1:] if row))) or dataset
        # 通常前四列是 dataset/version/metric/mode，后面是模型列。
        if len(header) >= 5:
            model = header[4]
        elif header:
            model = header[-1]

    safe = f"{model}-{dataset}-{input_dir.name}.xlsx"
    return re.sub(r"[\\/:*?\"<>|]+", "_", safe)


def write_summary_sheet(ws, md_text: str) -> None:
    ws.title = "summary"
    table = parse_markdown_table(md_text)

    if table:
        for row in table:
            ws.append(row)
    else:
        for line in md_text.splitlines():
            ws.append([line])

    max_row = max(ws.max_row, 1)
    max_col = max(ws.max_column, 1)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in range(1, max_col + 1):
        width = 16
        for row in range(1, min(max_row, 50) + 1):
            value = ws.cell(row, col).value
            if value is not None:
                width = max(width, min(len(str(value)) + 2, 32))
        ws.column_dimensions[get_column_letter(col)].width = width


def write_details_sheet(ws, rows: List[List[Any]]) -> None:
    ws.title = "details"
    headers = ["题号", "是否正确", "模型结果", "原始题目", "标准答案", "回答是否为空"]
    ws.append(headers)
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ok_fill = PatternFill("solid", fgColor="E2F0D9")
    bad_fill = PatternFill("solid", fgColor="FCE4D6")
    empty_fill = PatternFill("solid", fgColor="FFF2CC")

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font

    for r in range(2, ws.max_row + 1):
        correct = str(ws.cell(r, 2).value).strip().upper() == "TRUE"
        empty = str(ws.cell(r, 6).value).strip().upper() == "TRUE"
        ws.cell(r, 2).fill = ok_fill if correct else bad_fill
        if empty:
            ws.cell(r, 6).fill = empty_fill

    widths = {
        "A": 10,   # 题号
        "B": 12,   # 是否正确
        "C": 70,   # 模型结果
        "D": 70,   # 原始题目
        "E": 24,   # 标准答案
        "F": 14,   # 回答是否为空
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # 长文本行给一点高度，但不搞成巨型瀑布。
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 90
    ws.row_dimensions[1].height = 24

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def build_detail_rows(input_dir: Path) -> Tuple[List[List[Any]], List[str]]:
    all_rows: List[List[Any]] = []
    datasets: List[str] = []

    result_files = find_result_files(input_dir)
    multiple_results = len(result_files) > 1

    for result_file in result_files:
        result_json = read_json(result_file)
        details = normalize_result_details(result_json)
        dataset = extract_dataset_name(result_file, result_json)
        datasets.append(dataset)

        pred_file = find_prediction_file(input_dir, result_file)
        pred_json = read_json(pred_file) if pred_file else None
        pred_by_key, pred_rows = normalize_prediction_map(pred_json)

        for i, detail in enumerate(details):
            pred_record = pick_prediction_record(detail, i, pred_by_key, pred_rows)
            raw_prediction = scalar(pred_record.get("prediction", ""))
            if is_empty_answer(raw_prediction):
                # 兜底：没有 prediction 原文时，用 result json 的抽取结果。
                raw_prediction = scalar(detail.get("pred", ""))

            question = extract_current_question(pred_record.get("origin_prompt", ""))
            standard_answer = scalar(detail.get("answer", ""))
            if is_empty_answer(standard_answer):
                standard_answer = scalar(pred_record.get("gold", ""))

            correct = to_bool(detail.get("correct", False))
            empty = is_empty_answer(raw_prediction)
            number = index_from_abbr(detail.get("example_abbr"))
            if number is None:
                number = str(i)

            # 如果一次跑多个数据集，为避免题号撞车，题号带 dataset 前缀。
            display_number = f"{dataset}_{number}" if multiple_results else number

            all_rows.append([
                display_number,
                bool(correct),
                str(raw_prediction),
                question,
                str(standard_answer),
                bool(empty),
            ])

    return all_rows, datasets


def export_to_excel(input_dir: Path, output_path: Optional[Path]) -> Path:
    input_dir = input_dir.resolve()
    if not input_dir.exists() or not input_dir.is_dir():
        die(f"input 不是有效文件夹：{input_dir}")

    summary_path = find_summary_md(input_dir)
    summary_md = summary_path.read_text(encoding="utf-8")
    detail_rows, datasets = build_detail_rows(input_dir)

    if output_path is None:
        output_path = input_dir / infer_default_output_name(summary_md, input_dir, datasets)
    else:
        output_path = output_path.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws_summary = wb.active
    write_summary_sheet(ws_summary, summary_md)
    ws_details = wb.create_sheet("details")
    write_details_sheet(ws_details, detail_rows)

    # 打开时默认停在 summary。
    wb.active = 0
    wb.save(output_path)
    return output_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export OpenCompass result/prediction/summary files to Excel.")
    parser.add_argument("--input", "-i", default=None, help="OpenCompass 单次运行结果文件夹；为空则用 outputs/default 下最新文件夹")
    parser.add_argument("--output", "-o", default=None, help="输出 .xlsx；为空则自动命名并放到 input 文件夹下")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_dir = Path(args.input) if args.input else latest_dir(Path("outputs") / "default")
    output_path = Path(args.output) if args.output else None
    written = export_to_excel(input_dir, output_path)
    print(f"Excel 已生成：{written}")


if __name__ == "__main__":
    main()
