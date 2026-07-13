#!/usr/bin/env python3
"""Run VLMEvalKit and export selected result columns, or process existing outputs.

Examples
--------
Run evaluation and export known columns:

    python tools/autoverse.py \
        --run \
        --target split Overall AR CP \
        -- \
        --model qwen3.5-35b-a3b \
        --base-url https://qianfan.baidubce.com/v2 \
        --key "$QIANFAN_KEY" \
        --data MMBench_V11_MINI \
        --mode all \
        --work-dir ./outputs/qwen_demo

Run evaluation, then select columns interactively:

    python tools/autoverse.py \
        --run \
        --interactive \
        -- \
        --model qwen3.5-35b-a3b \
        --base-url https://qianfan.baidubce.com/v2 \
        --key "$QIANFAN_KEY" \
        --data MMBench_V11_MINI \
        --mode all \
        --work-dir ./outputs/qwen_demo

Process an existing result directory:

    python tools/autoverse.py \
        --input ./outputs/qwen_demo \
        --target split Overall AR CP

Inspect files and columns without exporting:

    python tools/autoverse.py \
        --input ./outputs/qwen_demo \
        --list-only
"""

from __future__ import annotations

import argparse
import hashlib
import shlex
import subprocess
import sys
from pathlib import Path
from typing import Iterable

import pandas as pd


SUPPORTED_SUFFIXES = {".csv", ".xlsx", ".xls"}
EXCLUDED_NAME_PARTS = {
    "_selected",
    "_checkpoint",
    "_prev",
    "_structs",
}


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Run VLMEvalKit or inspect an existing result directory, then "
            "export selected columns to a new Excel file."
        )
    )

    mode_group = parser.add_mutually_exclusive_group(required=True)
    mode_group.add_argument(
        "--run",
        action="store_true",
        help="Run the original VLMEvalKit run.py first.",
    )
    mode_group.add_argument(
        "--input",
        help="Existing VLMEvalKit result file or directory.",
    )

    parser.add_argument(
        "--run-script",
        default="run.py",
        help="Path to the original VLMEvalKit run.py. Default: run.py",
    )
    parser.add_argument(
        "--target",
        nargs="+",
        default=None,
        help="Target result columns to export.",
    )
    parser.add_argument(
        "--interactive",
        action="store_true",
        help="Select result file and columns interactively.",
    )
    parser.add_argument(
        "--file",
        default=None,
        help="Optional full or partial source filename.",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output .xlsx path. Default: <source>_selected.xlsx",
    )
    parser.add_argument(
        "--list-only",
        action="store_true",
        help="Only list result files and legal columns; do not export.",
    )

    return parser


def split_autoverse_and_run_args(argv: list[str]) -> tuple[list[str], list[str]]:
    """Split arguments at `--`; everything after it belongs to run.py."""
    if "--" not in argv:
        return argv, []

    index = argv.index("--")
    return argv[:index], argv[index + 1:]


def parse_args() -> tuple[argparse.Namespace, list[str]]:
    own_args, run_args = split_autoverse_and_run_args(sys.argv[1:])
    args = build_parser().parse_args(own_args)

    if args.run and not run_args:
        raise ValueError(
            "--run requires VLMEvalKit arguments after `--`.\n"
            "Example: autoverse.py --run --target Overall -- "
            "--model MODEL --data DATASET"
        )

    if not args.list_only and not args.interactive and not args.target:
        raise ValueError(
            "Choose one output mode: --target COLUMN... or --interactive."
        )

    return args, run_args


def extract_option_value(arguments: list[str], option: str) -> str | None:
    """Read a single-value option from run.py arguments."""
    for index, argument in enumerate(arguments):
        if argument == option and index + 1 < len(arguments):
            return arguments[index + 1]
        if argument.startswith(option + "="):
            return argument.split("=", 1)[1]
    return None


def run_vlmeval(run_script: Path, run_args: list[str]) -> Path:
    """Run the original VLMEvalKit command and return its work directory."""
    if not run_script.exists():
        raise FileNotFoundError(f"run.py not found: {run_script}")

    work_dir_value = extract_option_value(run_args, "--work-dir")
    work_dir = Path(work_dir_value or "./outputs").expanduser().resolve()

    command = [sys.executable, str(run_script), *run_args]

    print("\n正在调用原版 VLMEvalKit：")
    print(shlex.join(command))
    print()

    completed = subprocess.run(command, check=False)
    if completed.returncode != 0:
        raise RuntimeError(
            f"VLMEvalKit run.py exited with code {completed.returncode}."
        )

    return work_dir


def is_candidate(path: Path) -> bool:
    if not path.is_file():
        return False
    if path.suffix.lower() not in SUPPORTED_SUFFIXES:
        return False

    lower_name = path.name.lower()
    if any(part in lower_name for part in EXCLUDED_NAME_PARTS):
        return False

    return True


def candidate_priority(path: Path) -> tuple[int, float]:
    """Prefer summary/evaluation files over raw inference files."""
    name = path.name.lower()

    if "_acc.csv" in name:
        rank = 0
    elif "score" in name:
        rank = 1
    elif "result" in name:
        rank = 2
    elif "eval" in name:
        rank = 3
    else:
        rank = 4

    return rank, -path.stat().st_mtime


def file_digest(path: Path) -> str:
    """Return a content hash used to remove duplicated result files."""
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def path_preference(path: Path) -> tuple[int, int, float]:
    """Prefer non-timestamp copies, then shallower paths, then newer files."""
    has_timestamp_dir = any(
        part.startswith("T") and part[1:].isdigit()
        for part in path.parts
    )
    return (
        1 if has_timestamp_dir else 0,
        len(path.parts),
        -path.stat().st_mtime,
    )


def remove_duplicate_files(files: list[Path]) -> list[Path]:
    """Merge files with identical content and keep the preferred path."""
    by_digest: dict[str, Path] = {}

    for path in files:
        digest = file_digest(path)
        current = by_digest.get(digest)

        if current is None or path_preference(path) < path_preference(current):
            by_digest[digest] = path

    return list(by_digest.values())


def discover_files(input_path: Path, filename_filter: str | None) -> list[Path]:
    if input_path.is_file():
        if not is_candidate(input_path):
            raise ValueError(f"Unsupported result file: {input_path}")
        return [input_path]

    if not input_path.is_dir():
        raise FileNotFoundError(f"Input path does not exist: {input_path}")

    files = [path for path in input_path.rglob("*") if is_candidate(path)]

    if filename_filter:
        needle = filename_filter.lower()
        files = [path for path in files if needle in path.name.lower()]

    files = remove_duplicate_files(files)
    return sorted(files, key=candidate_priority)


def read_table(path: Path) -> pd.DataFrame:
    suffix = path.suffix.lower()

    if suffix == ".csv":
        return pd.read_csv(path)
    if suffix in {".xlsx", ".xls"}:
        return pd.read_excel(path)

    raise ValueError(f"Unsupported result type: {path.suffix}")


def print_candidate_files(files: Iterable[Path]) -> None:
    print("\n合法结果文件：")
    for index, path in enumerate(files, start=1):
        print(f"{index}. {path}")


def select_source_file(files: list[Path], interactive: bool) -> Path:
    if not files:
        raise FileNotFoundError("No legal CSV or Excel result file was found.")

    if len(files) == 1:
        return files[0]

    if not interactive:
        selected = files[0]
        print(f"\n自动选择优先级最高的结果文件：{selected}")
        return selected

    print_candidate_files(files)

    while True:
        raw = input("\n请选择结果文件编号：").strip()
        try:
            index = int(raw)
        except ValueError:
            print("请输入有效数字。")
            continue

        if 1 <= index <= len(files):
            return files[index - 1]

        print("编号超出范围。")


def select_source_for_targets(
    files: list[Path],
    requested: list[str],
) -> tuple[Path, pd.DataFrame]:
    """Choose the deduplicated result file that best matches target columns."""
    if not files:
        raise FileNotFoundError("No legal CSV or Excel result file was found.")

    requested = deduplicate(requested)
    ranked: list[tuple[int, int, tuple[int, float], Path, pd.DataFrame]] = []

    for path in files:
        try:
            frame = read_table(path)
        except Exception:
            continue

        columns = [str(column) for column in frame.columns]
        found_count = sum(column in columns for column in requested)
        all_found = int(found_count == len(requested))
        ranked.append(
            (
                -all_found,
                -found_count,
                candidate_priority(path),
                path,
                frame,
            )
        )

    if not ranked:
        raise ValueError("No readable result file was found.")

    ranked.sort(key=lambda item: (item[0], item[1], item[2]))
    _, _, _, source, frame = ranked[0]

    print(f"\n自动定位结果文件：{source}")
    return source, frame


def display_columns(columns: list[str]) -> None:
    print("\n合法列：")
    for index, column in enumerate(columns, start=1):
        print(f"{index}. {column}")


def parse_selection(raw: str, legal_columns: list[str]) -> list[str]:
    parts = [part.strip() for part in raw.split(",") if part.strip()]
    if not parts:
        raise ValueError("没有选择任何列。")

    selected: list[str] = []

    for part in parts:
        if part.isdigit():
            index = int(part)
            if not 1 <= index <= len(legal_columns):
                raise ValueError(f"列编号超出范围：{part}")
            column = legal_columns[index - 1]
        else:
            if part not in legal_columns:
                raise ValueError(f"不存在的列：{part}")
            column = part

        if column not in selected:
            selected.append(column)

    return selected


def interactive_select_columns(legal_columns: list[str]) -> list[str]:
    display_columns(legal_columns)

    while True:
        raw = input(
            "\n请输入列编号或列名，使用逗号分隔，例如 1,2,4 "
            "或 split,Overall,CP："
        ).strip()

        try:
            return parse_selection(raw, legal_columns)
        except ValueError as exc:
            print(f"选择无效：{exc}")


def deduplicate(columns: Iterable[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()

    for column in columns:
        if column not in seen:
            seen.add(column)
            result.append(column)

    return result


def repair_target_columns(
    requested: list[str],
    legal_columns: list[str],
) -> list[str]:
    """Validate targets and enter correction mode when some columns are absent."""
    requested = deduplicate(requested)
    found = [column for column in requested if column in legal_columns]
    missing = [column for column in requested if column not in legal_columns]

    if not missing:
        print(f"\nTarget validation passed: {found}")
        return found

    print("\n进入 Target 修正模式：")
    print(f"Found: {found if found else 'None'}")
    print(f"Not found: {missing}")
    display_columns(legal_columns)

    if not sys.stdin.isatty():
        if found:
            print("\n当前不是交互式终端，将只导出 Found 列。")
            return found
        raise ValueError("None of the requested target columns exists.")

    while True:
        print("\n请选择：")
        print("1. 直接输出 Found 列")
        print("2. 在 Found 列基础上添加其他合法列")
        print("3. 重新选择全部输出列")
        print("4. 取消")

        choice = input("> ").strip()

        if choice == "1":
            if not found:
                print("没有可输出的 Found 列，请选择 2 或 3。")
                continue
            return found

        if choice == "2":
            raw = input(
                "请输入要追加的列编号或列名，逗号分隔："
            ).strip()
            try:
                added = parse_selection(raw, legal_columns)
            except ValueError as exc:
                print(f"选择无效：{exc}")
                continue
            return deduplicate([*found, *added])

        if choice == "3":
            return interactive_select_columns(legal_columns)

        if choice == "4":
            raise KeyboardInterrupt

        print("请输入 1、2、3 或 4。")


def default_output_path(source: Path) -> Path:
    return source.with_name(f"{source.stem}_selected.xlsx")


def list_files_and_columns(files: list[Path]) -> None:
    if not files:
        print("没有找到合法结果文件。")
        return

    print_candidate_files(files)

    for path in files:
        print(f"\n文件：{path}")
        try:
            frame = read_table(path)
        except Exception as exc:
            print(f"读取失败：{exc}")
            continue
        display_columns([str(column) for column in frame.columns])


def export_columns(
    source: Path,
    frame: pd.DataFrame,
    columns: list[str],
    output: str | None,
) -> Path:
    """Export a readable Excel workbook with wrapping and sensible sizing."""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    output_path = (
        Path(output).expanduser().resolve()
        if output
        else default_output_path(source)
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)

    selected = frame.loc[:, columns].copy()
    selected.to_excel(output_path, index=False, sheet_name="Selected Results")

    workbook = load_workbook(output_path)
    sheet = workbook["Selected Results"]
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    for index, column in enumerate(sheet.columns, start=1):
        values = ["" if cell.value is None else str(cell.value) for cell in column]
        max_length = max((len(value) for value in values), default=0)

        header = str(sheet.cell(row=1, column=index).value or "").lower()
        if header in {"question", "hint", "prediction", "log"}:
            width = min(max(max_length * 0.8, 24), 60)
        else:
            width = min(max(max_length + 2, 10), 28)

        sheet.column_dimensions[get_column_letter(index)].width = width

    for row_index in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row_index].height = 45

    sheet.row_dimensions[1].height = 24
    workbook.save(output_path)
    return output_path


def main() -> int:
    try:
        args, run_args = parse_args()

        if args.run:
            run_script = Path(args.run_script).expanduser().resolve()
            input_path = run_vlmeval(run_script, run_args)
        else:
            input_path = Path(args.input).expanduser().resolve()

        files = discover_files(input_path, args.file)

        if args.list_only:
            list_files_and_columns(files)
            return 0

        if args.target is not None and not args.interactive:
            source, frame = select_source_for_targets(files, args.target)
        else:
            source = select_source_file(files, interactive=args.interactive)
            frame = read_table(source)

        legal_columns = [str(column) for column in frame.columns]
        print(f"\n当前结果文件：{source}")

        if args.interactive:
            selected_columns = interactive_select_columns(legal_columns)
        else:
            display_columns(legal_columns)
            selected_columns = repair_target_columns(
                requested=args.target,
                legal_columns=legal_columns,
            )

        output_path = export_columns(
            source=source,
            frame=frame,
            columns=selected_columns,
            output=args.output,
        )

        print("\n精准导出完成。")
        print(f"来源文件：{source}")
        print(f"输出列：{selected_columns}")
        print(f"输出文件：{output_path}")

        return 0

    except KeyboardInterrupt:
        print("\n操作已取消。")
        return 130
    except Exception as exc:
        print(f"错误：{exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
